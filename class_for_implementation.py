import os
import time
import numpy as np
import pandas as pd
import cv2
import tensorflow as tf
from tensorflow import keras
from keras.preprocessing.image import ImageDataGenerator
import zipfile
import shutil
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border
import glob
import pytesseract
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import pdb

'''
required:

-tensorflow
-cv2
-pytesseract
-keras
-fuzzywuzzy

***** input files are excel files with images

_init_ initilizates required functions

--'C:\Program Files\Tesseract-OCR' se necesita esta carpeta en esa ruta, esta en onebox en mi handover
'''
class Watermark_Detection:
    def _init_(self, source_path, destination_path):
        self.destination_path = destination_path
        self.path_output = destination_path
        self.source_path = source_path
        self.dic,self.dic_validation,self.dic_data = {},{},{}
        self.clean_source()
        self.check_if_path_is_folder()
        self.get_images_from_input_files()
        self.load_model()
        for i in [0,90,180,270]:
            self.get_processed_images(i)
            self.use_model()
            print(i)
        #start_time = time.time()
        self.read_image_text_to_datadic()
        #print("--- %s seconds ---" % (time.time() - start_time))
        self.output()

    def clean_source(self): #Function that cleans input and result forlder from source path
        try:
            shutil.rmtree(r''+self.destination_path +"\\results")
            print("Folder source limpio")
        except:
            print("Fallo al limpiar folder results") 
        try:
            shutil.rmtree(r''+self.destination_path+"\\input")
        except:
            print("Fallo al limpiar folder input")
        self.destination_path = self.destination_path + "\\input"
    
    def check_if_path_is_folder(self):#Funtion to check if path is folder or file and also revalue sourcepath
        try:
            os.makedirs(self.destination_path + '\\files')
            print("Folder Input creado")
        except:
            print('Folder Input ya existe')
        if os.path.isfile(self.source_path):
            try:
                shutil.copy(self.source_path, self.destination_path+ '\\files') #Copy file to input folder
                print("El path es archivo")
            except:
                print("No se copio el archivo")
            
        #Check if path is forder
        elif os.path.isdir(self.source_path):
            file_list = os.listdir(self.source_path)
            try:
                for name in file_list: #Copy files to input folder
                    shutil.copy(self.source_path + '\\'+name, self.destination_path+ '\\files')
                print("El path es folder")
            except:
                print("No se copiaron los archivos")
        self.source_path = self.destination_path + '\\files'

    def get_images_from_input_files(self): #Function to extract all image from inputs and put then on inputs/img folder
        print('in in get_images_from_input_files')
        #get names of the input files
        lst_input_file_names = os.listdir(self.source_path)
        image_counter = 1
        
        #Crear carpeta img
        path_img_folder = r''+self.destination_path+'\img'
        try:
            os.makedirs(path_img_folder)
        except:
            print('error o folder ya creado')

        #itera los archivos inputs... num = contador, file_name = nombre del archivo
        for num,file_name in enumerate(lst_input_file_names):
            sufix = "." + file_name.split(".")[-1]
            print(sufix)
            print(file_name.split(sufix)[0])
            self.dic_data[file_name.split(sufix)[0]] = {}
            self.dic_data[file_name.split(sufix)[0]]['tresh1'] = {}
            self.dic_data[file_name.split(sufix)[0]]['tresh2'] = {}
            self.dic_data[file_name.split(sufix)[0]]['tresh3'] = {}
            #Definiendo rutas y creando folders
            path_current_input_file = r'' + self.source_path + '/' + file_name
            path_current_input_folder = r''+self.destination_path+'\data\data'+str(num)
            if sufix == '.xlsx':
                path_current_input_media = path_current_input_folder + r'\xl\media'
            else:
                path_current_input_media = path_current_input_folder + r'\word\media'
            path_current_img_filename = r''+path_img_folder+'\\'+ file_name.split(sufix)[0]
            print(path_current_input_file)
            print(path_current_input_folder)
            print(path_current_input_media)
            print(path_current_img_filename)
            try:
                os.makedirs(path_current_input_folder)
            except:
                print('error o folder ya creado')
            try:
                os.makedirs(path_current_img_filename)
            except:
                print('error o folder ya creado')
            try:
                with zipfile.ZipFile(path_current_input_file,"r") as zip_ref: #unzip Excel file
                    zip_ref.extractall(path_current_input_folder)
                    print("Excel file unzip successfully at: ", path_current_input_folder)
            except OSError as e:
                print("Failed to create directory. Error:", e)
            
            #Rename and copy images to img folder
            num_images = 0
            for images in glob.iglob(os.path.join(path_current_input_media, "*")):
                rename_img = path_current_input_media + "\\" +"imagen"+str(image_counter)+".jpg" 
                image_counter = image_counter + 1 #counter to get a number fo all images
                os.rename(images,rename_img) #rename image
                num_images = num_images + 1
                shutil.copy(rename_img, path_current_img_filename) #copy image to img

    def get_processed_images(self,angle): #Function to process image form inputs/img with cv2
    
        #Definir rutas
        path_img_folder = r''+ self.destination_path+'\img'
        path_predict_folder = r''+self.destination_path + "\Predict"
        
        try:
            shutil.rmtree(path_predict_folder)
        except:
            print("Fallo al limpiar folder predict") 
        try:
            os.makedirs(path_predict_folder)
        except:
            print('error o folder predict ya creado')

        lst_file_name = os.listdir(path_img_folder)
        #itera los archivos del folder img
        #archivos_excel = [archivo for archivo in os.listdir(file_folder) if archivo.endswith('.xlsx')]
        for file_name in lst_file_name:
            path_img_filename = r''+path_img_folder+ '\\'+ file_name

            if len(os.listdir(path_img_filename)) == 0:
                print("Directory is empty")
                path_tresh1 = path_predict_folder + '\\tresh1\\'+ file_name
                path_tresh2 = path_predict_folder + '\\tresh2\\'+ file_name
                path_tresh3 = path_predict_folder + '\\tresh3\\'+ file_name

                try:
                    os.makedirs(path_tresh1)
                    os.makedirs(path_tresh2)
                    os.makedirs(path_tresh3)

                except:
                    print('error')
            else:
                print("Directory is not empty")
                salida = 1
                #Define rutas para la prediccion     #path_predict_folder = r''+destination_path + "\Predict"
                path_tresh1 = path_predict_folder + '\\tresh1\\'+ file_name
                path_tresh2 = path_predict_folder + '\\tresh2\\'+ file_name
                path_tresh3 = path_predict_folder + '\\tresh3\\'+ file_name

                try:
                    os.makedirs(path_tresh1)
                    os.makedirs(path_tresh2)
                    os.makedirs(path_tresh3)

                except:
                    print('error')

                #itera las imagenes del archivo
                for image in os.listdir(path_img_filename):
                    
                    #rota la imagen segun el angulo proporcionado
                    if angle == 0:
                        readed_img = cv2.imread(r''+path_img_filename +"\\"+image)
                    elif angle == 90:
                        readed_img = cv2.imread(r''+path_img_filename +"\\"+image)
                        readed_img= cv2.rotate(readed_img, cv2.ROTATE_90_CLOCKWISE)
                    elif angle == 180:
                        readed_img = cv2.imread(r''+path_img_filename +"\\"+image)
                        readed_img= cv2.rotate(readed_img, cv2.ROTATE_180)
                    elif angle == 270:
                        readed_img = cv2.imread(r''+path_img_filename +"\\"+image)
                        readed_img= cv2.rotate(readed_img, cv2.ROTATE_90_COUNTERCLOCKWISE)

                    height, width, c = readed_img.shape

                    #GRAYSCALE
                    im_gray = cv2.cvtColor(readed_img, cv2.COLOR_RGB2GRAY)

                    #THRESHOLD
                    (tresh, treshImage) = cv2.threshold(im_gray, 250, 255, cv2.THRESH_BINARY)
                    (tresh2, treshImage2) = cv2.threshold(im_gray, 220, 255, cv2.THRESH_BINARY)
                    (tresh3, treshImage3) = cv2.threshold(im_gray, 175, 255, cv2.THRESH_BINARY)

                    #Dividiendo imagen a la mitad
                    treshImage2 = treshImage2[ int(height/2):int(height), 0: width]
                    treshImage = treshImage[ int(height/2):int(height), 0: width]
                    treshImage3 = treshImage3[ int(height/2):int(height), 0: width]

                    #Guardando la img preprosesadas
                    cv2.imwrite(path_tresh1 + "\\" + image, treshImage2)
                    cv2.imwrite(path_tresh2 + "\\" + image, treshImage)
                    cv2.imwrite(path_tresh3 + "\\" + image, treshImage3)

    def find_string_ratio(self,watermark_word,data,r1,r2): #Function to find considenses in the string
            passed = 0
            ratio = fuzz.token_set_ratio(watermark_word.lower(),data.lower())
            partial_ratio = fuzz.partial_ratio(watermark_word.lower(),data.lower())
            if partial_ratio > r1 or ratio > r2:
                '''print(i)
                print("Partial ratio: ", partial_ratio)
                print("Ratio: ", ratio)
                print("\n\n")'''
                passed = 1
            return passed    
    
    def load_model(self): #Function that Loads the model from .h5 file and set img size
        
        self.test_datagen = ImageDataGenerator(rescale=1./255)
        
        self.batch_size = 32
        self.img_size = (320,400)

        #Carga el modelo
        #new_model = tf.keras.models.load_model('deeplearning.h5')
        self.new_model = tf.keras.models.load_model(r'C:\Program Files\Tesseract-OCR\deeplearning.h5')
        #self.new_model = tf.keras.models.load_model('deeplearning.h5')

    def use_model(self): #Apply model and outputs a dic with 1 or 0

        self.dic_img = {}
        self.tresh_lst = ['tresh1','tresh2','tresh3']
        for self.current_tresh in self.tresh_lst:
            self.path_current_tresh = self.destination_path + '\\Predict\\' + self.current_tresh
            file_list = os.listdir(self.path_current_tresh)
            for self.file_name in file_list:
                print("leyendo Archivo")
                path_current_filename = self.path_current_tresh + '\\'+self.file_name
                if len(os.listdir(path_current_filename)) > 0:
                    print("modelo utilizado en tresh: % s file name: % s " % (self.current_tresh,self.file_name))
                    it = self.test_datagen.flow_from_directory(self.path_current_tresh,
                        color_mode='rgb',
                        classes=[self.file_name],
                        shuffle=False,
                        target_size=self.img_size)
                    preds = self.new_model.predict_generator(it)
                    
                    lst = []
                    #si existe el nombre del archivo en el diccionario lo copia
                    if self.file_name in self.dic:
                        self.dic_img = self.dic[self.file_name]

                    #itera en la lista predicciones
                    for num,i in enumerate(preds):

                        if str(it.filenames[num]).split("\\")[1].split(".")[0] not in self.dic_img:
                            self.dic_img[str(it.filenames[num]).split("\\")[1].split(".")[0]] = 0

                        #Check if image is positive or not then save negative image on differnte folder.
                        if np.argmax(i) == 1:
                            self.dic_img[it.filenames[num].split("\\")[1].split(".")[0]] = 1
                    
                    self.copy_image_and_delete_from_source()
                    self.dic[self.file_name] = self.dic_img
                    self.dic_img = {}

                else:
                    self.dic_img = {}
                    print('No images')
    
    def copy_image_and_delete_from_source(self): #Function that copy images to /read/tresh  and deletes them on /img

        folder_to_copy = self.destination_path + '\\read' + '\\' + self.current_tresh + '\\'+ self.file_name
        folder_to_delete = self.destination_path + "\Predict"
        try:
            os.makedirs(folder_to_copy)
        except:
            print('carpeta ya creada')
        for image_name in self.dic_img:
            if self.dic_img[image_name] == 1:
                img_to_copy = self.path_current_tresh + '\\' + self.file_name + '\\'+ image_name+'.jpg'
                
                if os.path.isfile(img_to_copy):
                    try:
                        img_to_detele= self.destination_path + "\\img"+'\\'+self.file_name+'\\'+ image_name+'.jpg'
                        shutil.copy(img_to_copy,folder_to_copy)
                    except:
                        print('Error copy image to read')
                    
                    if os.path.isfile(img_to_detele):
                        try:   
                            os.remove(img_to_detele)
                        except:
                            print('error borrando img')

    def read_image_text_to_datadic(self):#Function that reads the text on the images and saves them on dic_data
        self.contador = 0
        self.suma = 0
        print('dic data keys', self.dic_data.keys())
        print('ejecutando read image')
        path_read_tresh1 = self.destination_path + '\\read\\tresh1'
        path_read_tresh2 = self.destination_path + '\\read\\tresh2'
        path_read_tresh3 = self.destination_path + '\\read\\tresh3'

        try:
            os.makedirs(path_read_tresh1)
            os.makedirs(path_read_tresh2)
            os.makedirs(path_read_tresh3)
        except:
            print('error creando folder')

        #self.dic_data = {}
        self.dic_img = {}

        for self.current_tresh in self.tresh_lst:
            path_class = self.destination_path + '\\read\\' + self.current_tresh
            print('pathclass ',path_class)
            file_list = os.listdir(path_class)
            print('file list: ',file_list)
            #itera las imagenes por archivs para predicciones
            for self.file_name in file_list:
                self.lst_current_img_name = os.listdir(path_class+ '\\' + self.file_name)
                
                if self.file_name in self.dic_data:
                    self.dic_img = self.dic_data[self.file_name][self.current_tresh]
                for img in self.lst_current_img_name:
                    if self.dic[self.file_name][img.split(".")[0]] >= 1:
                        loaded_img = cv2.imread(path_class + '\\' + self.file_name+'\\'+img)
                        try:    
                            data = pytesseract.image_to_string(loaded_img, lang='eng', config='--psm 6')
                            self.contador += 1
                            self.suma = len(data) + self.suma
                        except:
                            print('error al leer imagenes')
                        try:
                            self.dic_img[img.split(".")[0]] = str(data)
                        except:
                            print('error dic img')
                            
                    else:
                        self.dic_img[img.split(".")[0]] = str('No data')
                self.dic_data[self.file_name][self.current_tresh] = self.dic_img
                self.find_text_coincidences()
                self.dic_img = {}

    def find_text_coincidences(self):#Function that finds the coincidences on for the watermark
        print('ejecutando find text')
        for i in range(0,len(self.lst_current_img_name) -6):
            ratio1 = self.find_string_ratio(self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i].split(".")[0]],self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i+1].split(".")[0]],75,65)
            ratio2 = self.find_string_ratio(self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i].split(".")[0]],self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i+2].split(".")[0]],75,65)
            ratio3 = self.find_string_ratio(self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i].split(".")[0]],self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i+3].split(".")[0]],75,65)
            print('ratio 1')
            if ratio1 == 1 or ratio2 == 1 or ratio3 == 1:
                ratio1 = self.find_string_ratio(self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i].split(".")[0]],self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i+4].split(".")[0]],65,60)
                ratio2 = self.find_string_ratio(self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i].split(".")[0]],self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i+5].split(".")[0]],65,60)
                ratio3 = self.find_string_ratio(self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i].split(".")[0]],self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i+6].split(".")[0]],65,60)
                if ratio1 == 1 or ratio2 == 1 or ratio3 == 1:
                    print('ratio 2')
                    watermark = self.dic_data[self.file_name][self.current_tresh][self.lst_current_img_name[i].split(".")[0]]
                    #print('watermark')
                    #print(watermark)
                    for img in self.lst_current_img_name:
                        if self.dic[self.file_name][img.split(".")[0]] >= 1:               
                            if self.find_string_ratio(watermark,self.dic_data[self.file_name][self.current_tresh][img.split(".")[0]],50,40) == 1:
                                self.dic[self.file_name][img.split(".")[0]] = int(self.dic[self.file_name][img.split(".")[0]]) + 1
                            elif len(self.dic_data[self.file_name][self.current_tresh][img.split(".")[0]]) > 110 and self.find_string_ratio(watermark,self.dic_data[self.file_name][self.current_tresh][img.split(".")[0]],38,30):
                                self.dic[self.file_name][img.split(".")[0]] = int(self.dic[self.file_name][img.split(".")[0]]) + 0.5
                    
                    break
                else: print('no se encontro coincidencias ratio 2')
            else:
                print('no se encontro coincidencias ratio 1')
        #print(self.dic)
        #print('promedio')
        #print(self.suma / self.contador)

    def output(self):#Funtion that makes all the ouput process
        try:
            shutil.rmtree(self.destination_path+"\\data")
        except:
            print('error borrar carpetar data')
        
        self.get_images_from_input_files()
        
        for i in self.dic:
            num_no_watermark = 0
            num_watermark = 0
            num_doubt_watermark = 0
            for j in self.dic[i]:
                try: 
                    os.makedirs(r''+self.path_output +'\\results\\'+ i+'\\NoMarcaAgua')
                except:
                    print('')
                
                try:
                    os.makedirs(r''+self.path_output +'\\results\\'+ i+'\\Duda')
                except:
                    print('')
                
                try:
                    os.makedirs(r''+self.path_output +'\\results\\'+ i+'\\ConMarcaAgua')
                except:
                    print('')

                if self.dic[i][j] <= 0: #Si imagen = 0 copia la foto a el foder No watermark
                    
                    num_no_watermark = num_no_watermark + 1
                    shutil.copy(r''+ self.destination_path+"\\img\\"+i+'\\'+j+'.jpg', r''+self.path_output +'\\results\\'+ i+'\\NoMarcaAgua')

                elif self.dic[i][j] <= 1.5: #Si imagen != 0 copia la foto a el foder Watermark
                    
                    num_doubt_watermark = num_doubt_watermark + 1
                    shutil.copy(r''+self.destination_path+"\\img\\"+i+'\\'+j+'.jpg', r''+self.path_output +'\\results\\'+ i+'\\Duda')
                
                else: #Si imagen != 0 copia la foto a el foder Watermark
                    
                    num_watermark = num_watermark + 1
                    shutil.copy(r''+self.destination_path+"\\img\\"+i+'\\'+j+'.jpg', r''+self.path_output +'\\results\\'+ i+'\\ConMarcaAgua')
            
            self.dic[i]['num_watermark'] = num_watermark
            self.dic[i]['num_no_watermark'] = num_no_watermark
            self.dic[i]['num_doubt_watermark'] = num_doubt_watermark
            self.dic[i]['total'] = num_no_watermark + num_watermark + num_doubt_watermark

        self.create_excel()

    def create_excel(self): #Create Output Excel
        
        wb= openpyxl.load_workbook(r'C:\Program Files\Tesseract-OCR\Template.xlsx')
        #wb= openpyxl.load_workbook(r'Template.xlsx')
        ws = wb.worksheets[0]
        lst = list(self.dic.keys())
        print('Final')
        for rows in range(2, len(self.dic)+2): #first element of rage refers on wich row do you want to start
            data = [lst[rows-2],self.dic[lst[rows-2]]['total'],self.dic[lst[rows-2]]['num_no_watermark'],self.dic[lst[rows-2]]['num_doubt_watermark'],self.dic[lst[rows-2]]['num_watermark']]
            for col in range(1,6):
                ws.cell(row=rows, column=col, value=data[col-1])
        #wb.save(r''+destination_path +'\\results\\Prototipo1.xlsx')
        wb.save(r''+self.path_output +'\\results\\resultados.xlsx')

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def main(path,results_path):
    Watermark_Detection(path,results_path)

#main(r'C:\Users\d84316956\Desktop\test_qc',r'C:\Users\d84316956\Desktop\New folder')
#send_email()